from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

def criar_modelo_importacao_tipos_lote():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tipos de Lote"
    
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:B1')
    ws['A1'] = 'INFORMAÇÕES BÁSICAS'
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('C1:G1')
    ws['C1'] = 'PREÇOS - LEVE (R$/kg)'
    ws['C1'].fill = header_fill
    ws['C1'].font = header_font
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('H1:L1')
    ws['H1'] = 'PREÇOS - MÉDIO (R$/kg)'
    ws['H1'].fill = header_fill
    ws['H1'].font = header_font
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('M1:Q1')
    ws['M1'] = 'PREÇOS - PESADO (R$/kg)'
    ws['M1'].fill = header_fill
    ws['M1'].font = header_font
    ws['M1'].alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['Nome', 'Descrição']
    for i in range(1, 6):
        headers.append(f'{i} Estrela{"s" if i > 1 else ""}')
    for i in range(1, 6):
        headers.append(f'{i} Estrela{"s" if i > 1 else ""}')
    for i in range(1, 6):
        headers.append(f'{i} Estrela{"s" if i > 1 else ""}')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        if col_num <= 2:
            cell.fill = subheader_fill
        else:
            cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    exemplo_data = [
        ['PET Transparente', 'Garrafas PET transparentes', 2.50, 2.80, 3.10, 3.40, 3.70, 
         2.20, 2.50, 2.80, 3.10, 3.40, 1.90, 2.20, 2.50, 2.80, 3.10],
        ['Papelão', 'Papelão ondulado limpo', 0.80, 1.00, 1.20, 1.40, 1.60,
         0.60, 0.80, 1.00, 1.20, 1.40, 0.40, 0.60, 0.80, 1.00, 1.20]
    ]
    
    for row_num, data in enumerate(exemplo_data, 3):
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num <= 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    for col in range(3, 18):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    
    instrucoes_ws = wb.create_sheet("Instruções")
    instrucoes_ws['A1'] = 'INSTRUÇÕES PARA IMPORTAÇÃO DE TIPOS DE LOTE'
    instrucoes_ws['A1'].font = Font(bold=True, size=14, color="059669")
    
    instrucoes = [
        '',
        '1. INFORMAÇÕES BÁSICAS:',
        '   - Nome: Nome do tipo de lote (obrigatório, único)',
        '   - Descrição: Descrição detalhada do tipo de lote (opcional)',
        '',
        '2. PREÇOS:',
        '   - Preencha os preços em R$ por kg para cada combinação de classificação e estrelas',
        '   - Leve: Material de fácil processamento',
        '   - Médio: Material de processamento padrão',
        '   - Pesado: Material de difícil processamento',
        '   - Estrelas: De 1 a 5, representando a qualidade do material',
        '',
        '3. OBSERVAÇÕES:',
        '   - As linhas 3 e 4 da planilha "Tipos de Lote" contêm exemplos',
        '   - Apague os exemplos antes de importar seus dados',
        '   - Você pode adicionar quantas linhas precisar',
        '   - O código do tipo de lote será gerado automaticamente',
        '   - Todos os preços devem ser números positivos',
        '   - Deixe em branco se não quiser configurar um preço específico',
        '',
        '4. IMPORTAÇÃO:',
        '   - Vá em Administração > Gerenciar Tipos de Lote',
        '   - Clique em "Importar Excel"',
        '   - Selecione o arquivo preenchido',
        '   - Aguarde a confirmação da importação',
    ]
    
    for row_num, instrucao in enumerate(instrucoes, 1):
        cell = instrucoes_ws.cell(row=row_num, column=1)
        cell.value = instrucao
        if row_num == 1:
            cell.font = Font(bold=True, size=14, color="059669")
        elif instrucao.startswith(('1.', '2.', '3.', '4.')):
            cell.font = Font(bold=True, size=11)
    
    instrucoes_ws.column_dimensions['A'].width = 80
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
