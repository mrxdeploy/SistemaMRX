from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.models import db, MaterialBase, TabelaPreco, TabelaPrecoItem, Usuario
from app.auth import admin_required
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
import logging # Import the logging module

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def garantir_tabelas_preco():
    """Garante que as 3 tabelas de preço necessárias existam no banco de dados"""
    tabelas_existentes = TabelaPreco.query.all()

    if len(tabelas_existentes) >= 3:
        return tabelas_existentes

    # Criar as tabelas faltantes
    niveis_necessarios = {1, 2, 3}
    niveis_existentes = {t.nivel_estrelas for t in tabelas_existentes}
    niveis_faltantes = niveis_necessarios - niveis_existentes

    for nivel in niveis_faltantes:
        nome_estrelas = "Estrela" if nivel == 1 else "Estrelas"
        tabela = TabelaPreco(
            nome=f'{nivel} {nome_estrelas}',
            nivel_estrelas=nivel,
            ativo=True
        )
        db.session.add(tabela)

    db.session.flush()
    return TabelaPreco.query.all()

bp = Blueprint('materiais_base', __name__, url_prefix='/api/materiais-base')

def gerar_codigo_automatico():
    ultimo_material = MaterialBase.query.order_by(MaterialBase.id.desc()).first()
    if ultimo_material and ultimo_material.codigo:
        try:
            numero = int(ultimo_material.codigo.replace('MAT', ''))
            return f'MAT{numero + 1:03d}'
        except:
            pass

    proximo_id = MaterialBase.query.count() + 1
    return f'MAT{proximo_id:03d}'

@bp.route('', methods=['GET'])
@jwt_required()
def listar_materiais():
    """Lista todos os materiais base ativos - SEM informações de preços para compradores"""
    try:
        usuario_id = get_jwt_identity()
        usuario = Usuario.query.get(usuario_id)

        materiais = MaterialBase.query.filter_by(ativo=True).order_by(MaterialBase.nome).all()

        # Se não for admin, remover informações de preços/estrelas
        if usuario and usuario.tipo != 'admin':
            return jsonify([{
                'id': m.id,
                'codigo': m.codigo,
                'nome': m.nome,
                'classificacao': m.classificacao,
                'descricao': m.descricao,
                'ativo': m.ativo,
                'data_cadastro': m.data_cadastro.isoformat() if m.data_cadastro else None,
                'data_atualizacao': m.data_atualizacao.isoformat() if m.data_atualizacao else None
            } for m in materiais]), 200

        # Admin vê tudo
        return jsonify([m.to_dict() for m in materiais]), 200
    except Exception as e:
        logger.error(f'Erro ao listar materiais: {str(e)}')
        return jsonify({'erro': f'Erro ao listar materiais: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def obter_material(id):
    try:
        material = MaterialBase.query.get(id)

        if not material:
            return jsonify({'erro': 'Material não encontrado'}), 404

        return jsonify(material.to_dict()), 200

    except Exception as e:
        logger.error(f'Erro ao obter material: {str(e)}')
        return jsonify({'erro': f'Erro ao obter material: {str(e)}'}), 500

@bp.route('', methods=['POST'])
@admin_required
def criar_material():
    try:
        data = request.get_json()

        required_fields = ['nome', 'classificacao']
        for field in required_fields:
            if field not in data:
                return jsonify({'erro': f'Campo {field} é obrigatório'}), 400

        if data['classificacao'] not in ['leve', 'medio', 'pesado']:
            return jsonify({'erro': 'Classificação deve ser: leve, medio ou pesado'}), 400

        material_existente = MaterialBase.query.filter_by(nome=data['nome']).first()
        if material_existente:
            return jsonify({'erro': 'Já existe um material com este nome'}), 400

        codigo = data.get('codigo') or gerar_codigo_automatico()
        codigo_existente = MaterialBase.query.filter_by(codigo=codigo).first()
        if codigo_existente:
            return jsonify({'erro': 'Código já está em uso'}), 400

        total_materiais = MaterialBase.query.count()
        if total_materiais >= 500:
            return jsonify({'erro': 'Limite de 500 materiais atingido'}), 400

        material = MaterialBase(
            codigo=codigo,
            nome=data['nome'],
            classificacao=data['classificacao'],
            descricao=data.get('descricao', ''),
            ativo=True
        )
        db.session.add(material)
        db.session.flush()

        tabelas_preco = garantir_tabelas_preco()
        precos = data.get('precos', {})

        if not precos:
            # Ao criar material, o fornecedor deve adicionar os materiais e valores corretamente
            # Portanto, não é mais obrigatório fornecer os preços no cadastro inicial.
            # O fornecedor adicionará os preços posteriormente na tabela de preços do fornecedor.
            pass
        else:
            # Validar que TODOS os 3 preços estão presentes e são válidos
            precos_validados = {}
            for tabela in tabelas_preco:
                preco_key = f'preco_{tabela.nivel_estrelas}_estrela'
                preco_valor_raw = precos.get(preco_key)

                # Rejeitar se o preço não foi fornecido
                if preco_valor_raw is None or preco_valor_raw == '':
                    return jsonify({'erro': f'Preço para {tabela.nivel_estrelas} estrela(s) é obrigatório'}), 400

                # Tentar converter para float
                try:
                    preco_valor = float(preco_valor_raw)
                except (ValueError, TypeError):
                    return jsonify({'erro': f'Preço para {tabela.nivel_estrelas} estrela(s) deve ser um número válido'}), 400

                # Validar que o preço é positivo
                if preco_valor <= 0:
                    return jsonify({'erro': f'Preço para {tabela.nivel_estrelas} estrela(s) deve ser maior que zero'}), 400

                precos_validados[tabela.id] = preco_valor

            # Agora criar os itens de preço com valores validados
            for tabela in tabelas_preco:
                preco_item = TabelaPrecoItem(
                    tabela_preco_id=tabela.id,
                    material_id=material.id,
                    preco_por_kg=precos_validados[tabela.id],
                    ativo=True
                )
                db.session.add(preco_item)

        db.session.commit()

        return jsonify({
            'mensagem': 'Material criado com sucesso',
            'material': material.to_dict()
        }), 201

    except Exception as e:
        db.session.rollback()
        logger.error(f'Erro ao criar material: {str(e)}')
        return jsonify({'erro': f'Erro ao criar material: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['PUT'])
@admin_required
def atualizar_material(id):
    try:
        material = MaterialBase.query.get(id)

        if not material:
            return jsonify({'erro': 'Material não encontrado'}), 404

        data = request.get_json()

        if 'nome' in data and data['nome'] != material.nome:
            material_existente = MaterialBase.query.filter_by(nome=data['nome']).first()
            if material_existente:
                return jsonify({'erro': 'Já existe um material com este nome'}), 400
            material.nome = data['nome']

        if 'classificacao' in data:
            if data['classificacao'] not in ['leve', 'medio', 'pesado']:
                return jsonify({'erro': 'Classificação deve ser: leve, medio ou pesado'}), 400
            material.classificacao = data['classificacao']

        if 'descricao' in data:
            material.descricao = data['descricao']

        if 'ativo' in data:
            material.ativo = data['ativo']

        if 'precos' in data:
            precos = data['precos']
            for preco_item in material.precos:
                tabela = preco_item.tabela_preco
                if tabela:
                    preco_key = f'preco_{tabela.nivel_estrelas}_estrela'
                    if preco_key in precos:
                        preco_valor_raw = precos[preco_key]

                        # Rejeitar valores inválidos
                        if preco_valor_raw is None or preco_valor_raw == '':
                            return jsonify({'erro': f'Preço para {tabela.nivel_estrelas} estrela(s) não pode estar vazio'}), 400

                        try:
                            preco_valor = float(preco_valor_raw)
                        except (ValueError, TypeError):
                            return jsonify({'erro': f'Preço para {tabela.nivel_estrelas} estrela(s) deve ser um número válido'}), 400

                        if preco_valor <= 0:
                            return jsonify({'erro': f'Preço para {tabela.nivel_estrelas} estrela(s) deve ser maior que zero'}), 400

                        preco_item.preco_por_kg = preco_valor

        material.data_atualizacao = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'mensagem': 'Material atualizado com sucesso',
            'material': material.to_dict()
        }), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f'Erro ao atualizar material: {str(e)}')
        return jsonify({'erro': f'Erro ao atualizar material: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['DELETE'])
@admin_required
def deletar_material(id):
    try:
        material = MaterialBase.query.get(id)

        if not material:
            return jsonify({'erro': 'Material não encontrado'}), 404

        material.ativo = False
        db.session.commit()

        return jsonify({'mensagem': 'Material desativado com sucesso'}), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f'Erro ao deletar material: {str(e)}')
        return jsonify({'erro': f'Erro ao deletar material: {str(e)}'}), 500

@bp.route('/importar-excel', methods=['POST'])
@admin_required
def importar_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'erro': 'Nome de arquivo inválido'}), 400

        df = pd.read_excel(file)

        # As colunas de preços agora são específicas para o cadastro de fornecedores
        # e não mais para a base geral de materiais.
        colunas_esperadas = ['Nome do Material', 'Classificação', 'Descrição']
        colunas_faltando = [col for col in colunas_esperadas if col not in df.columns]

        if colunas_faltando:
            return jsonify({'erro': f'Colunas faltando no Excel: {", ".join(colunas_faltando)}'}), 400

        tabelas_preco = {tab.nivel_estrelas: tab for tab in TabelaPreco.query.all()}

        if len(tabelas_preco) < 3:
            return jsonify({'erro': 'É necessário ter as 3 tabelas de preço cadastradas'}), 400

        materiais_criados = 0
        materiais_atualizados = 0
        erros = []

        for index, row in df.iterrows():
            try:
                nome = str(row['Nome do Material']).strip()
                classificacao = str(row['Classificação']).strip().lower()
                descricao = str(row.get('Descrição', '')).strip()

                if classificacao not in ['leve', 'medio', 'pesado']:
                    erros.append(f"Linha {index+2}: Classificação inválida '{classificacao}'")
                    continue

                material = MaterialBase.query.filter_by(nome=nome).first()

                if not material:
                    codigo = gerar_codigo_automatico()
                    material = MaterialBase(
                        codigo=codigo,
                        nome=nome,
                        classificacao=classificacao,
                        descricao=descricao,
                        ativo=True
                    )
                    db.session.add(material)
                    db.session.flush()
                    materiais_criados += 1
                else:
                    # Atualiza apenas a classificação e descrição se o material já existe
                    material.classificacao = classificacao
                    material.descricao = descricao
                    materiais_atualizados += 1

            except Exception as e:
                erros.append(f"Linha {index+2}: {str(e)}")

        db.session.commit()

        return jsonify({
            'mensagem': 'Importação concluída',
            'materiais_criados': materiais_criados,
            'materiais_atualizados': materiais_atualizados,
            'erros': erros
        }), 200

    except Exception as e:
        db.session.rollback()
        logger.error(f'Erro ao importar Excel: {str(e)}')
        return jsonify({'erro': f'Erro ao importar Excel: {str(e)}'}), 500

@bp.route('/exportar-excel', methods=['GET'])
@admin_required
def exportar_excel():
    try:
        materiais = MaterialBase.query.order_by(MaterialBase.codigo).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Materiais'

        headers = ['Código', 'Nome do Material', 'Classificação', 'Descrição']
        ws.append(headers)

        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for material in materiais:
            row_data = [
                material.codigo,
                material.nome,
                material.classificacao.capitalize(),
                material.descricao or ''
            ]
            ws.append(row_data)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'materiais_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        logger.error(f'Erro ao exportar Excel: {str(e)}')
        return jsonify({'erro': f'Erro ao exportar Excel: {str(e)}'}), 500

@bp.route('/modelo-importacao', methods=['GET'])
@admin_required
def modelo_importacao():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Materiais'

        headers = ['Nome do Material', 'Classificação', 'Descrição']
        ws.append(headers)

        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        exemplos = [
            ['SUCATA PROCESSADOR CERÂMICO A', 'pesado', 'Processador cerâmico tipo A'],
            ['SUCATA PLACA MÃE SERVIDOR', 'pesado', 'Placa mãe de servidor'],
            ['SUCATA HD SATA', 'medio', 'HD SATA'],
        ]

        for exemplo in exemplos:
            ws.append(exemplo)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='modelo_importacao_materiais.xlsx'
        )

    except Exception as e:
        logger.error(f'Erro ao gerar modelo: {str(e)}')
        return jsonify({'erro': f'Erro ao gerar modelo: {str(e)}'}), 500