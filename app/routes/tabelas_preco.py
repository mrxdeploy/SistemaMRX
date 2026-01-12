from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from app.models import db, TabelaPreco, TabelaPrecoItem, MaterialBase
from app.auth import admin_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

bp = Blueprint('tabelas_preco', __name__, url_prefix='/api/tabelas-preco')

@bp.route('', methods=['GET'])
@jwt_required()
def listar_tabelas():
    try:
        tabelas = TabelaPreco.query.order_by(TabelaPreco.nivel_estrelas).all()
        return jsonify([tabela.to_dict() for tabela in tabelas]), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao listar tabelas: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def obter_tabela(id):
    try:
        tabela = TabelaPreco.query.get(id)
        
        if not tabela:
            return jsonify({'erro': 'Tabela não encontrada'}), 404
        
        return jsonify(tabela.to_dict()), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao obter tabela: {str(e)}'}), 500

@bp.route('/<int:id>/precos', methods=['GET'])
@jwt_required()
def listar_precos(id):
    try:
        tabela = TabelaPreco.query.get(id)
        
        if not tabela:
            return jsonify({'erro': 'Tabela não encontrada'}), 404
        
        busca = request.args.get('busca', '')
        classificacao = request.args.get('classificacao', '')
        apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'
        
        query = TabelaPrecoItem.query.filter_by(tabela_preco_id=id)
        
        if apenas_ativos:
            query = query.filter_by(ativo=True)
        
        if busca:
            query = query.join(MaterialBase).filter(
                db.or_(
                    MaterialBase.nome.ilike(f'%{busca}%'),
                    MaterialBase.codigo.ilike(f'%{busca}%')
                )
            )
        
        if classificacao and classificacao in ['leve', 'medio', 'pesado']:
            query = query.join(MaterialBase).filter(MaterialBase.classificacao == classificacao)
        
        precos = query.order_by(MaterialBase.codigo).all()
        
        return jsonify([preco.to_dict() for preco in precos]), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao listar preços: {str(e)}'}), 500

@bp.route('/<int:tabela_id>/precos/<int:material_id>', methods=['PUT'])
@admin_required
def atualizar_preco_individual(tabela_id, material_id):
    try:
        tabela = TabelaPreco.query.get(tabela_id)
        if not tabela:
            return jsonify({'erro': 'Tabela não encontrada'}), 404
        
        material = MaterialBase.query.get(material_id)
        if not material:
            return jsonify({'erro': 'Material não encontrado'}), 404
        
        data = request.get_json()
        
        if 'preco_por_kg' not in data:
            return jsonify({'erro': 'Campo preco_por_kg é obrigatório'}), 400
        
        preco_por_kg = float(data['preco_por_kg'])
        
        if preco_por_kg < 0:
            return jsonify({'erro': 'Preço não pode ser negativo'}), 400
        
        preco_item = TabelaPrecoItem.query.filter_by(
            tabela_preco_id=tabela_id,
            material_id=material_id
        ).first()
        
        if not preco_item:
            preco_item = TabelaPrecoItem(
                tabela_preco_id=tabela_id,
                material_id=material_id,
                preco_por_kg=preco_por_kg,
                ativo=True
            )
            db.session.add(preco_item)
        else:
            preco_item.preco_por_kg = preco_por_kg
            preco_item.data_atualizacao = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Preço atualizado com sucesso',
            'preco': preco_item.to_dict()
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao atualizar preço: {str(e)}'}), 500

@bp.route('/<int:id>/precos', methods=['PUT'])
@admin_required
def atualizar_precos_massa(id):
    try:
        tabela = TabelaPreco.query.get(id)
        
        if not tabela:
            return jsonify({'erro': 'Tabela não encontrada'}), 404
        
        data = request.get_json()
        
        if 'precos' not in data or not isinstance(data['precos'], list):
            return jsonify({'erro': 'Campo precos (lista) é obrigatório'}), 400
        
        atualizados = 0
        erros = []
        
        for item in data['precos']:
            try:
                material_id = item.get('material_id')
                preco_por_kg = float(item.get('preco_por_kg', 0.00))
                
                if not material_id:
                    continue
                
                if preco_por_kg < 0:
                    erros.append(f"Material {material_id}: preço não pode ser negativo")
                    continue
                
                preco_item = TabelaPrecoItem.query.filter_by(
                    tabela_preco_id=id,
                    material_id=material_id
                ).first()
                
                if preco_item:
                    preco_item.preco_por_kg = preco_por_kg
                    preco_item.data_atualizacao = datetime.utcnow()
                    atualizados += 1
            
            except Exception as e:
                erros.append(f"Material {item.get('material_id', 'N/A')}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Preços atualizados',
            'atualizados': atualizados,
            'erros': erros
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao atualizar preços: {str(e)}'}), 500

@bp.route('/<int:id>/importar-excel', methods=['POST'])
@admin_required
def importar_precos_excel(id):
    try:
        tabela = TabelaPreco.query.get(id)
        
        if not tabela:
            return jsonify({'erro': 'Tabela não encontrada'}), 404
        
        if 'file' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'erro': 'Nome de arquivo inválido'}), 400
        
        import pandas as pd
        df = pd.read_excel(file)
        
        colunas_esperadas = ['Código', 'Nome do Material', 'Preço (R$/kg)']
        colunas_faltando = [col for col in colunas_esperadas if col not in df.columns]
        
        if colunas_faltando:
            return jsonify({'erro': f'Colunas faltando no Excel: {", ".join(colunas_faltando)}'}), 400
        
        precos_atualizados = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                codigo = str(row['Código']).strip()
                preco_valor = float(row['Preço (R$/kg)'])
                
                material = MaterialBase.query.filter_by(codigo=codigo).first()
                
                if not material:
                    erros.append(f"Linha {index+2}: Material {codigo} não encontrado")
                    continue
                
                if preco_valor < 0:
                    erros.append(f"Linha {index+2}: Preço não pode ser negativo")
                    continue
                
                preco_item = TabelaPrecoItem.query.filter_by(
                    tabela_preco_id=id,
                    material_id=material.id
                ).first()
                
                if not preco_item:
                    preco_item = TabelaPrecoItem(
                        tabela_preco_id=id,
                        material_id=material.id,
                        preco_por_kg=preco_valor,
                        ativo=True
                    )
                    db.session.add(preco_item)
                else:
                    preco_item.preco_por_kg = preco_valor
                    preco_item.data_atualizacao = datetime.utcnow()
                
                precos_atualizados += 1
            
            except Exception as e:
                erros.append(f"Linha {index+2}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Importação concluída',
            'precos_atualizados': precos_atualizados,
            'erros': erros
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao importar Excel: {str(e)}'}), 500

@bp.route('/<int:id>/exportar-excel', methods=['GET'])
@admin_required
def exportar_precos_excel(id):
    try:
        tabela = TabelaPreco.query.get(id)
        
        if not tabela:
            return jsonify({'erro': 'Tabela não encontrada'}), 404
        
        precos = TabelaPrecoItem.query.filter_by(tabela_preco_id=id).order_by(MaterialBase.codigo).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Preços {tabela.nivel_estrelas} Estrela'
        
        headers = ['Código', 'Nome do Material', 'Classificação', 'Preço (R$/kg)']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for preco_item in precos:
            material = preco_item.material
            if material:
                row_data = [
                    material.codigo,
                    material.nome,
                    material.classificacao.capitalize(),
                    float(preco_item.preco_por_kg)
                ]
                ws.append(row_data)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'precos_{tabela.nivel_estrelas}_estrela_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao exportar Excel: {str(e)}'}), 500
