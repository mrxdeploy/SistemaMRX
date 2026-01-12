from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from app.models import db, FornecedorTipoLoteClassificacao, Fornecedor, TipoLote
from app.auth import admin_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import pandas as pd
from datetime import datetime

bp = Blueprint('fornecedor_tipo_lote_classificacoes', __name__, url_prefix='/api/fornecedor-tipo-lote-classificacoes')

@bp.route('', methods=['GET'])
@jwt_required()
def listar_classificacoes():
    try:
        fornecedor_id = request.args.get('fornecedor_id', type=int)
        tipo_lote_id = request.args.get('tipo_lote_id', type=int)
        apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'
        
        query = FornecedorTipoLoteClassificacao.query
        
        if apenas_ativos:
            query = query.filter_by(ativo=True)
        
        if fornecedor_id:
            query = query.filter_by(fornecedor_id=fornecedor_id)
        
        if tipo_lote_id:
            query = query.filter_by(tipo_lote_id=tipo_lote_id)
        
        classificacoes = query.all()
        return jsonify([c.to_dict() for c in classificacoes]), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao listar classificações: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def obter_classificacao(id):
    try:
        classificacao = FornecedorTipoLoteClassificacao.query.get(id)
        
        if not classificacao:
            return jsonify({'erro': 'Classificação não encontrada'}), 404
        
        return jsonify(classificacao.to_dict()), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao obter classificação: {str(e)}'}), 500

@bp.route('', methods=['POST'])
@admin_required
def criar_classificacao():
    try:
        data = request.get_json()
        
        if not data or not data.get('fornecedor_id') or not data.get('tipo_lote_id'):
            return jsonify({'erro': 'fornecedor_id e tipo_lote_id são obrigatórios'}), 400
        
        fornecedor = Fornecedor.query.get(data['fornecedor_id'])
        if not fornecedor:
            return jsonify({'erro': 'Fornecedor não encontrado'}), 404
        
        tipo_lote = TipoLote.query.get(data['tipo_lote_id'])
        if not tipo_lote:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        classificacao_existente = FornecedorTipoLoteClassificacao.query.filter_by(
            fornecedor_id=data['fornecedor_id'],
            tipo_lote_id=data['tipo_lote_id']
        ).first()
        
        if classificacao_existente:
            return jsonify({'erro': 'Já existe uma classificação para este fornecedor e tipo de lote'}), 400
        
        for campo in ['leve_estrelas', 'medio_estrelas', 'pesado_estrelas']:
            valor = data.get(campo, 1 if campo == 'leve_estrelas' else (3 if campo == 'medio_estrelas' else 5))
            if valor < 1 or valor > 5:
                return jsonify({'erro': f'{campo} deve estar entre 1 e 5'}), 400
        
        classificacao = FornecedorTipoLoteClassificacao(
            fornecedor_id=data['fornecedor_id'],
            tipo_lote_id=data['tipo_lote_id'],
            leve_estrelas=data.get('leve_estrelas', 1),
            medio_estrelas=data.get('medio_estrelas', 3),
            pesado_estrelas=data.get('pesado_estrelas', 5),
            ativo=data.get('ativo', True)
        )
        
        db.session.add(classificacao)
        db.session.commit()
        
        return jsonify(classificacao.to_dict()), 201
    
    except ValueError as ve:
        db.session.rollback()
        return jsonify({'erro': str(ve)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao criar classificação: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['PUT'])
@admin_required
def atualizar_classificacao(id):
    try:
        classificacao = FornecedorTipoLoteClassificacao.query.get(id)
        
        if not classificacao:
            return jsonify({'erro': 'Classificação não encontrada'}), 404
        
        data = request.get_json()
        
        if 'fornecedor_id' in data and data['fornecedor_id'] != classificacao.fornecedor_id:
            fornecedor = Fornecedor.query.get(data['fornecedor_id'])
            if not fornecedor:
                return jsonify({'erro': 'Fornecedor não encontrado'}), 404
            classificacao.fornecedor_id = data['fornecedor_id']
        
        if 'tipo_lote_id' in data and data['tipo_lote_id'] != classificacao.tipo_lote_id:
            tipo_lote = TipoLote.query.get(data['tipo_lote_id'])
            if not tipo_lote:
                return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
            classificacao.tipo_lote_id = data['tipo_lote_id']
        
        for campo in ['leve_estrelas', 'medio_estrelas', 'pesado_estrelas']:
            if campo in data:
                valor = data[campo]
                if valor < 1 or valor > 5:
                    return jsonify({'erro': f'{campo} deve estar entre 1 e 5'}), 400
                setattr(classificacao, campo, valor)
        
        if 'ativo' in data:
            classificacao.ativo = data['ativo']
        
        db.session.commit()
        
        return jsonify(classificacao.to_dict()), 200
    
    except ValueError as ve:
        db.session.rollback()
        return jsonify({'erro': str(ve)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao atualizar classificação: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['DELETE'])
@admin_required
def deletar_classificacao(id):
    try:
        classificacao = FornecedorTipoLoteClassificacao.query.get(id)
        
        if not classificacao:
            return jsonify({'erro': 'Classificação não encontrada'}), 404
        
        db.session.delete(classificacao)
        db.session.commit()
        
        return jsonify({'mensagem': 'Classificação deletada com sucesso'}), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao deletar classificação: {str(e)}'}), 500

@bp.route('/modelo-excel', methods=['GET'])
@jwt_required()
def download_modelo_excel():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modelo de Importação"
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            'Fornecedor ID',
            'Nome Fornecedor',
            'Tipo Lote ID',
            'Nome Tipo Lote',
            'Estrelas Leve (1-5)',
            'Estrelas Médio (1-5)',
            'Estrelas Pesado (1-5)',
            'Ativo (SIM/NÃO)'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        fornecedores = Fornecedor.query.filter_by(ativo=True).order_by(Fornecedor.nome).all()
        tipos_lote = TipoLote.query.filter_by(ativo=True).order_by(TipoLote.nome).all()
        
        row = 2
        for fornecedor in fornecedores[:5]:
            for tipo_lote in tipos_lote[:3]:
                ws.cell(row=row, column=1, value=fornecedor.id)
                ws.cell(row=row, column=2, value=fornecedor.nome)
                ws.cell(row=row, column=3, value=tipo_lote.id)
                ws.cell(row=row, column=4, value=tipo_lote.nome)
                ws.cell(row=row, column=5, value=1)
                ws.cell(row=row, column=6, value=3)
                ws.cell(row=row, column=7, value=5)
                ws.cell(row=row, column=8, value='SIM')
                row += 1
        
        ws_ref = wb.create_sheet("Fornecedores Disponíveis")
        ws_ref.append(['ID', 'Nome', 'CNPJ/CPF'])
        for fornecedor in fornecedores:
            ws_ref.append([
                fornecedor.id,
                fornecedor.nome,
                fornecedor.cnpj or fornecedor.cpf or ''
            ])
        
        ws_ref2 = wb.create_sheet("Tipos de Lote Disponíveis")
        ws_ref2.append(['ID', 'Nome', 'Código'])
        for tipo in tipos_lote:
            ws_ref2.append([
                tipo.id,
                tipo.nome,
                tipo.codigo or ''
            ])
        
        for ws_temp in [ws, ws_ref, ws_ref2]:
            for column_cells in ws_temp.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws_temp.column_dimensions[column_cells[0].column_letter].width = min(50, length + 2)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'modelo_classificacoes_estrelas_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao gerar modelo Excel: {str(e)}'}), 500

@bp.route('/importar-excel', methods=['POST'])
@admin_required
def importar_excel():
    try:
        if 'arquivo' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['arquivo']
        
        if file.filename == '':
            return jsonify({'erro': 'Nenhum arquivo selecionado'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'erro': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400
        
        df = pd.read_excel(file, sheet_name=0)
        
        required_columns = [
            'Fornecedor ID',
            'Tipo Lote ID',
            'Estrelas Leve (1-5)',
            'Estrelas Médio (1-5)',
            'Estrelas Pesado (1-5)'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'erro': f'Colunas obrigatórias faltando: {", ".join(missing_columns)}'}), 400
        
        df = df.dropna(subset=['Fornecedor ID', 'Tipo Lote ID'])
        
        criados = 0
        atualizados = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                fornecedor_id = int(row['Fornecedor ID'])
                tipo_lote_id = int(row['Tipo Lote ID'])
                leve = int(row['Estrelas Leve (1-5)'])
                medio = int(row['Estrelas Médio (1-5)'])
                pesado = int(row['Estrelas Pesado (1-5)'])
                
                if leve < 1 or leve > 5 or medio < 1 or medio > 5 or pesado < 1 or pesado > 5:
                    erros.append(f'Linha {index + 2}: Estrelas devem estar entre 1 e 5')
                    continue
                
                fornecedor = Fornecedor.query.get(fornecedor_id)
                if not fornecedor:
                    erros.append(f'Linha {index + 2}: Fornecedor ID {fornecedor_id} não encontrado')
                    continue
                
                tipo_lote = TipoLote.query.get(tipo_lote_id)
                if not tipo_lote:
                    erros.append(f'Linha {index + 2}: Tipo de Lote ID {tipo_lote_id} não encontrado')
                    continue
                
                ativo_valor = str(row.get('Ativo (SIM/NÃO)', 'SIM')).strip().upper()
                ativo = ativo_valor in ['SIM', 'S', 'YES', 'Y', '1', 'TRUE']
                
                classificacao = FornecedorTipoLoteClassificacao.query.filter_by(
                    fornecedor_id=fornecedor_id,
                    tipo_lote_id=tipo_lote_id
                ).first()
                
                if classificacao:
                    classificacao.leve_estrelas = leve
                    classificacao.medio_estrelas = medio
                    classificacao.pesado_estrelas = pesado
                    classificacao.ativo = ativo
                    atualizados += 1
                else:
                    nova_classificacao = FornecedorTipoLoteClassificacao(
                        fornecedor_id=fornecedor_id,
                        tipo_lote_id=tipo_lote_id,
                        leve_estrelas=leve,
                        medio_estrelas=medio,
                        pesado_estrelas=pesado,
                        ativo=ativo
                    )
                    db.session.add(nova_classificacao)
                    criados += 1
            
            except ValueError as ve:
                erros.append(f'Linha {index + 2}: {str(ve)}')
            except Exception as e:
                erros.append(f'Linha {index + 2}: Erro ao processar - {str(e)}')
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Importação concluída',
            'criados': criados,
            'atualizados': atualizados,
            'erros': erros,
            'total_erros': len(erros)
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao importar arquivo Excel: {str(e)}'}), 500

@bp.route('/exportar-excel', methods=['GET'])
@jwt_required()
def exportar_excel():
    try:
        apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'
        
        query = FornecedorTipoLoteClassificacao.query
        
        if apenas_ativos:
            query = query.filter_by(ativo=True)
        
        classificacoes = query.all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Classificações de Estrelas"
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            'ID',
            'Fornecedor ID',
            'Nome Fornecedor',
            'Tipo Lote ID',
            'Nome Tipo Lote',
            'Estrelas Leve (1-5)',
            'Estrelas Médio (1-5)',
            'Estrelas Pesado (1-5)',
            'Ativo',
            'Data Cadastro',
            'Data Atualização'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row_idx, classificacao in enumerate(classificacoes, start=2):
            ws.cell(row=row_idx, column=1, value=classificacao.id)
            ws.cell(row=row_idx, column=2, value=classificacao.fornecedor_id)
            ws.cell(row=row_idx, column=3, value=classificacao.fornecedor.nome if classificacao.fornecedor else '')
            ws.cell(row=row_idx, column=4, value=classificacao.tipo_lote_id)
            ws.cell(row=row_idx, column=5, value=classificacao.tipo_lote.nome if classificacao.tipo_lote else '')
            ws.cell(row=row_idx, column=6, value=classificacao.leve_estrelas)
            ws.cell(row=row_idx, column=7, value=classificacao.medio_estrelas)
            ws.cell(row=row_idx, column=8, value=classificacao.pesado_estrelas)
            ws.cell(row=row_idx, column=9, value='SIM' if classificacao.ativo else 'NÃO')
            ws.cell(row=row_idx, column=10, value=classificacao.data_cadastro.strftime('%d/%m/%Y %H:%M') if classificacao.data_cadastro else '')
            ws.cell(row=row_idx, column=11, value=classificacao.data_atualizacao.strftime('%d/%m/%Y %H:%M') if classificacao.data_atualizacao else '')
        
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(50, length + 2)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'classificacoes_estrelas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao exportar para Excel: {str(e)}'}), 500
