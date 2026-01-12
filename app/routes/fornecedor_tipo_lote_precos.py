from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from app.models import db, FornecedorTipoLotePreco, Fornecedor, TipoLote
from app.auth import admin_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import pandas as pd
from datetime import datetime

bp = Blueprint('fornecedor_tipo_lote_precos', __name__, url_prefix='/api/fornecedor-tipo-lote-precos')

@bp.route('', methods=['GET'])
@jwt_required()
def listar_precos():
    try:
        fornecedor_id = request.args.get('fornecedor_id', type=int)
        tipo_lote_id = request.args.get('tipo_lote_id', type=int)
        apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'
        
        query = FornecedorTipoLotePreco.query
        
        if apenas_ativos:
            query = query.filter_by(ativo=True)
        
        if fornecedor_id:
            query = query.filter_by(fornecedor_id=fornecedor_id)
        
        if tipo_lote_id:
            query = query.filter_by(tipo_lote_id=tipo_lote_id)
        
        precos = query.all()
        return jsonify([p.to_dict() for p in precos]), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao listar preços: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def obter_preco(id):
    try:
        preco = FornecedorTipoLotePreco.query.get(id)
        
        if not preco:
            return jsonify({'erro': 'Preço não encontrado'}), 404
        
        return jsonify(preco.to_dict()), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao obter preço: {str(e)}'}), 500

@bp.route('', methods=['POST'])
@admin_required
def criar_preco():
    try:
        data = request.get_json()
        
        required_fields = ['fornecedor_id', 'tipo_lote_id', 'estrelas', 'preco_por_kg']
        for field in required_fields:
            if field not in data:
                return jsonify({'erro': f'Campo {field} é obrigatório'}), 400
        
        fornecedor = Fornecedor.query.get(data['fornecedor_id'])
        if not fornecedor:
            return jsonify({'erro': 'Fornecedor não encontrado'}), 404
        
        tipo_lote = TipoLote.query.get(data['tipo_lote_id'])
        if not tipo_lote:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        if data['estrelas'] < 1 or data['estrelas'] > 5:
            return jsonify({'erro': 'Estrelas deve estar entre 1 e 5'}), 400
        
        if data['preco_por_kg'] < 0:
            return jsonify({'erro': 'Preço por kg deve ser maior ou igual a zero'}), 400
        
        preco_existente = FornecedorTipoLotePreco.query.filter_by(
            fornecedor_id=data['fornecedor_id'],
            tipo_lote_id=data['tipo_lote_id'],
            estrelas=data['estrelas']
        ).first()
        
        if preco_existente:
            return jsonify({'erro': 'Já existe um preço para este fornecedor, tipo de lote e quantidade de estrelas'}), 400
        
        preco = FornecedorTipoLotePreco(
            fornecedor_id=data['fornecedor_id'],
            tipo_lote_id=data['tipo_lote_id'],
            estrelas=data['estrelas'],
            preco_por_kg=data['preco_por_kg'],
            ativo=data.get('ativo', True)
        )
        
        db.session.add(preco)
        db.session.commit()
        
        return jsonify(preco.to_dict()), 201
    
    except ValueError as ve:
        db.session.rollback()
        return jsonify({'erro': str(ve)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao criar preço: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['PUT'])
@admin_required
def atualizar_preco(id):
    try:
        preco = FornecedorTipoLotePreco.query.get(id)
        
        if not preco:
            return jsonify({'erro': 'Preço não encontrado'}), 404
        
        data = request.get_json()
        
        novo_fornecedor_id = data.get('fornecedor_id', preco.fornecedor_id)
        novo_tipo_lote_id = data.get('tipo_lote_id', preco.tipo_lote_id)
        novas_estrelas = data.get('estrelas', preco.estrelas)
        
        if 'fornecedor_id' in data and data['fornecedor_id'] != preco.fornecedor_id:
            fornecedor = Fornecedor.query.get(data['fornecedor_id'])
            if not fornecedor:
                return jsonify({'erro': 'Fornecedor não encontrado'}), 404
        
        if 'tipo_lote_id' in data and data['tipo_lote_id'] != preco.tipo_lote_id:
            tipo_lote = TipoLote.query.get(data['tipo_lote_id'])
            if not tipo_lote:
                return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        if 'estrelas' in data:
            if data['estrelas'] < 1 or data['estrelas'] > 5:
                return jsonify({'erro': 'Estrelas deve estar entre 1 e 5'}), 400
        
        if (novo_fornecedor_id != preco.fornecedor_id or 
            novo_tipo_lote_id != preco.tipo_lote_id or 
            novas_estrelas != preco.estrelas):
            
            duplicata = FornecedorTipoLotePreco.query.filter_by(
                fornecedor_id=novo_fornecedor_id,
                tipo_lote_id=novo_tipo_lote_id,
                estrelas=novas_estrelas
            ).filter(FornecedorTipoLotePreco.id != id).first()
            
            if duplicata:
                return jsonify({'erro': 'Já existe um preço para este fornecedor, tipo de lote e quantidade de estrelas'}), 400
        
        preco.fornecedor_id = novo_fornecedor_id
        preco.tipo_lote_id = novo_tipo_lote_id
        preco.estrelas = novas_estrelas
        
        if 'preco_por_kg' in data:
            if data['preco_por_kg'] < 0:
                return jsonify({'erro': 'Preço por kg deve ser maior ou igual a zero'}), 400
            preco.preco_por_kg = data['preco_por_kg']
        
        if 'ativo' in data:
            preco.ativo = data['ativo']
        
        db.session.commit()
        
        return jsonify(preco.to_dict()), 200
    
    except ValueError as ve:
        db.session.rollback()
        return jsonify({'erro': str(ve)}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao atualizar preço: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['DELETE'])
@admin_required
def deletar_preco(id):
    try:
        preco = FornecedorTipoLotePreco.query.get(id)
        
        if not preco:
            return jsonify({'erro': 'Preço não encontrado'}), 404
        
        db.session.delete(preco)
        db.session.commit()
        
        return jsonify({'mensagem': 'Preço deletado com sucesso'}), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao deletar preço: {str(e)}'}), 500

@bp.route('/lote/<int:fornecedor_id>/<int:tipo_lote_id>', methods=['POST'])
@admin_required
def criar_precos_lote(fornecedor_id, tipo_lote_id):
    """Cria múltiplos preços (1-5 estrelas) de uma vez para um fornecedor/tipo"""
    try:
        data = request.get_json()
        
        fornecedor = Fornecedor.query.get(fornecedor_id)
        if not fornecedor:
            return jsonify({'erro': 'Fornecedor não encontrado'}), 404
        
        tipo_lote = TipoLote.query.get(tipo_lote_id)
        if not tipo_lote:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        if 'precos' not in data or not isinstance(data['precos'], dict):
            return jsonify({'erro': 'Campo precos é obrigatório e deve ser um dicionário {estrelas: valor}'}), 400
        
        criados = []
        atualizados = []
        
        for estrelas_str, preco_valor in data['precos'].items():
            estrelas = int(estrelas_str)
            
            if estrelas < 1 or estrelas > 5:
                continue
            
            if preco_valor < 0:
                continue
            
            preco_existente = FornecedorTipoLotePreco.query.filter_by(
                fornecedor_id=fornecedor_id,
                tipo_lote_id=tipo_lote_id,
                estrelas=estrelas
            ).first()
            
            if preco_existente:
                preco_existente.preco_por_kg = preco_valor
                preco_existente.ativo = True
                atualizados.append(preco_existente.to_dict())
            else:
                novo_preco = FornecedorTipoLotePreco(
                    fornecedor_id=fornecedor_id,
                    tipo_lote_id=tipo_lote_id,
                    estrelas=estrelas,
                    preco_por_kg=preco_valor,
                    ativo=True
                )
                db.session.add(novo_preco)
                criados.append({'estrelas': estrelas, 'preco_por_kg': preco_valor})
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Preços salvos com sucesso',
            'criados': len(criados),
            'atualizados': len(atualizados)
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao criar preços em lote: {str(e)}'}), 500

@bp.route('/modelo-excel', methods=['GET'])
@jwt_required()
def download_modelo_excel():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modelo Preços por Estrela"
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            'Fornecedor ID',
            'Nome Fornecedor',
            'Tipo Lote ID',
            'Nome Tipo Lote',
            'Estrelas (1-5)',
            'Preço por KG (R$)',
            'Ativo (SIM/NÃO)'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        fornecedores = Fornecedor.query.filter_by(ativo=True).order_by(Fornecedor.nome).limit(3).all()
        tipos_lote = TipoLote.query.filter_by(ativo=True).order_by(TipoLote.nome).limit(2).all()
        
        row = 2
        for fornecedor in fornecedores:
            for tipo_lote in tipos_lote:
                for estrelas in range(1, 6):
                    ws.cell(row=row, column=1, value=fornecedor.id)
                    ws.cell(row=row, column=2, value=fornecedor.nome)
                    ws.cell(row=row, column=3, value=tipo_lote.id)
                    ws.cell(row=row, column=4, value=tipo_lote.nome)
                    ws.cell(row=row, column=5, value=estrelas)
                    ws.cell(row=row, column=6, value=round(10.00 + (estrelas * 2.5), 2))
                    ws.cell(row=row, column=7, value='SIM')
                    row += 1
        
        ws_ref = wb.create_sheet("Fornecedores Disponíveis")
        ws_ref.append(['ID', 'Nome', 'CNPJ/CPF'])
        for fornecedor in Fornecedor.query.filter_by(ativo=True).all():
            ws_ref.append([
                fornecedor.id,
                fornecedor.nome,
                fornecedor.cnpj or fornecedor.cpf or ''
            ])
        
        ws_ref2 = wb.create_sheet("Tipos de Lote Disponíveis")
        ws_ref2.append(['ID', 'Nome', 'Código'])
        for tipo in TipoLote.query.filter_by(ativo=True).all():
            ws_ref2.append([
                tipo.id,
                tipo.nome,
                tipo.codigo or ''
            ])
        
        for ws_temp in [ws, ws_ref, ws_ref2]:
            for column_cells in ws_temp.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws_temp.column_dimensions[column_cells[0].column_letter].width = min(50, length + 2)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'modelo_precos_estrelas_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao gerar modelo Excel: {str(e)}'}), 500

@bp.route('/importar-excel', methods=['POST'])
@admin_required
def importar_excel():
    try:
        if 'arquivo' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['arquivo']
        
        if file.filename == '':
            return jsonify({'erro': 'Nenhum arquivo selecionado'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'erro': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400
        
        df = pd.read_excel(file, sheet_name=0)
        
        required_columns = [
            'Fornecedor ID',
            'Tipo Lote ID',
            'Estrelas (1-5)',
            'Preço por KG (R$)'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'erro': f'Colunas obrigatórias faltando: {", ".join(missing_columns)}'}), 400
        
        df = df.dropna(subset=['Fornecedor ID', 'Tipo Lote ID', 'Estrelas (1-5)', 'Preço por KG (R$)'])
        
        criados = 0
        atualizados = 0
        erros = []
        
        for index, row in df.iterrows():
            try:
                fornecedor_id = int(row['Fornecedor ID'])
                tipo_lote_id = int(row['Tipo Lote ID'])
                estrelas = int(row['Estrelas (1-5)'])
                preco_kg = float(row['Preço por KG (R$)'])
                
                if estrelas < 1 or estrelas > 5:
                    erros.append(f'Linha {index + 2}: Estrelas deve estar entre 1 e 5')
                    continue
                
                if preco_kg < 0:
                    erros.append(f'Linha {index + 2}: Preço por kg deve ser maior ou igual a zero')
                    continue
                
                fornecedor = Fornecedor.query.get(fornecedor_id)
                if not fornecedor:
                    erros.append(f'Linha {index + 2}: Fornecedor ID {fornecedor_id} não encontrado')
                    continue
                
                tipo_lote = TipoLote.query.get(tipo_lote_id)
                if not tipo_lote:
                    erros.append(f'Linha {index + 2}: Tipo de Lote ID {tipo_lote_id} não encontrado')
                    continue
                
                ativo_valor = str(row.get('Ativo (SIM/NÃO)', 'SIM')).strip().upper()
                ativo = ativo_valor in ['SIM', 'S', 'YES', 'Y', '1', 'TRUE']
                
                preco = FornecedorTipoLotePreco.query.filter_by(
                    fornecedor_id=fornecedor_id,
                    tipo_lote_id=tipo_lote_id,
                    estrelas=estrelas
                ).first()
                
                if preco:
                    preco.preco_por_kg = preco_kg
                    preco.ativo = ativo
                    atualizados += 1
                else:
                    novo_preco = FornecedorTipoLotePreco(
                        fornecedor_id=fornecedor_id,
                        tipo_lote_id=tipo_lote_id,
                        estrelas=estrelas,
                        preco_por_kg=preco_kg,
                        ativo=ativo
                    )
                    db.session.add(novo_preco)
                    criados += 1
            
            except ValueError as ve:
                erros.append(f'Linha {index + 2}: {str(ve)}')
            except Exception as e:
                erros.append(f'Linha {index + 2}: Erro ao processar - {str(e)}')
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Importação concluída',
            'criados': criados,
            'atualizados': atualizados,
            'erros': erros,
            'total_erros': len(erros)
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao importar arquivo Excel: {str(e)}'}), 500

@bp.route('/exportar-excel', methods=['GET'])
@jwt_required()
def exportar_excel():
    try:
        apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'
        fornecedor_id = request.args.get('fornecedor_id', type=int)
        
        query = FornecedorTipoLotePreco.query
        
        if apenas_ativos:
            query = query.filter_by(ativo=True)
        
        if fornecedor_id:
            query = query.filter_by(fornecedor_id=fornecedor_id)
        
        precos = query.all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Preços por Estrela"
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            'ID',
            'Fornecedor ID',
            'Nome Fornecedor',
            'Tipo Lote ID',
            'Nome Tipo Lote',
            'Estrelas',
            'Preço por KG (R$)',
            'Ativo',
            'Data Cadastro',
            'Data Atualização'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row_idx, preco in enumerate(precos, start=2):
            ws.cell(row=row_idx, column=1, value=preco.id)
            ws.cell(row=row_idx, column=2, value=preco.fornecedor_id)
            ws.cell(row=row_idx, column=3, value=preco.fornecedor.nome if preco.fornecedor else '')
            ws.cell(row=row_idx, column=4, value=preco.tipo_lote_id)
            ws.cell(row=row_idx, column=5, value=preco.tipo_lote.nome if preco.tipo_lote else '')
            ws.cell(row=row_idx, column=6, value=preco.estrelas)
            ws.cell(row=row_idx, column=7, value=preco.preco_por_kg)
            ws.cell(row=row_idx, column=8, value='SIM' if preco.ativo else 'NÃO')
            ws.cell(row=row_idx, column=9, value=preco.data_cadastro.strftime('%d/%m/%Y %H:%M') if preco.data_cadastro else '')
            ws.cell(row=row_idx, column=10, value=preco.data_atualizacao.strftime('%d/%m/%Y %H:%M') if preco.data_atualizacao else '')
        
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(50, length + 2)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'precos_estrelas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao exportar para Excel: {str(e)}'}), 500
