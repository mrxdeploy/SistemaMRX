from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from app.models import TipoLote, TipoLotePreco, db, FornecedorTipoLoteClassificacao, Fornecedor
from app.auth import admin_required
from app.utils.excel_template import criar_modelo_importacao_tipos_lote
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

bp = Blueprint('tipos_lote', __name__, url_prefix='/api/tipos-lote')

def gerar_codigo_automatico():
    ultimo_tipo = TipoLote.query.order_by(TipoLote.id.desc()).first()
    if ultimo_tipo and ultimo_tipo.codigo:
        try:
            numero = int(ultimo_tipo.codigo.replace('TL', ''))
            return f'TL{numero + 1:03d}'
        except:
            pass
    
    proximo_id = TipoLote.query.count() + 1
    return f'TL{proximo_id:03d}'

@bp.route('', methods=['GET'])
@jwt_required()
def listar_tipos_lote():
    try:
        busca = request.args.get('busca', '')
        apenas_ativos = request.args.get('apenas_ativos', 'true').lower() == 'true'
        
        query = TipoLote.query
        
        if apenas_ativos:
            query = query.filter_by(ativo=True)
        
        if busca:
            query = query.filter(
                db.or_(
                    TipoLote.nome.ilike(f'%{busca}%'),
                    TipoLote.codigo.ilike(f'%{busca}%'),
                    TipoLote.descricao.ilike(f'%{busca}%')
                )
            )
        
        tipos = query.order_by(TipoLote.nome).all()
        return jsonify([tipo.to_dict() for tipo in tipos]), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao listar tipos de lote: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def obter_tipo_lote(id):
    try:
        tipo = TipoLote.query.get(id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        return jsonify(tipo.to_dict()), 200
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao obter tipo de lote: {str(e)}'}), 500

@bp.route('', methods=['POST'])
@admin_required
def criar_tipo_lote():
    try:
        data = request.get_json()
        
        if not data or not data.get('nome'):
            return jsonify({'erro': 'Nome é obrigatório'}), 400
        
        tipo_existente = TipoLote.query.filter_by(nome=data['nome']).first()
        if tipo_existente:
            return jsonify({'erro': 'Já existe um tipo de lote com este nome'}), 400
        
        codigo = data.get('codigo', '').strip()
        if not codigo:
            codigo = gerar_codigo_automatico()
        else:
            codigo_existente = TipoLote.query.filter_by(codigo=codigo).first()
            if codigo_existente:
                return jsonify({'erro': 'Já existe um tipo de lote com este código'}), 400
        
        total_tipos = TipoLote.query.count()
        if total_tipos >= 150:
            return jsonify({'erro': 'Limite máximo de 150 tipos de lote atingido'}), 400
        
        classificacao = data.get('classificacao', None)
        if classificacao and classificacao not in ['leve', 'media', 'pesada', '']:
            return jsonify({'erro': 'Classificação deve ser: leve, media ou pesada'}), 400
        
        if classificacao == '':
            classificacao = None
        
        tipo = TipoLote(
            nome=data['nome'],
            descricao=data.get('descricao', ''),
            codigo=codigo,
            classificacao=classificacao,
            ativo=data.get('ativo', True)
        )
        
        db.session.add(tipo)
        db.session.flush()
        
        if 'precos' in data and isinstance(data['precos'], dict):
            for classificacao_key in ['leve', 'medio', 'pesado']:
                if classificacao_key in data['precos']:
                    precos_class = data['precos'][classificacao_key]
                    for estrela in range(1, 6):
                        preco_val = precos_class.get(str(estrela), 0.0)
                        if isinstance(preco_val, (int, float)) and preco_val >= 0:
                            preco_obj = TipoLotePreco(
                                tipo_lote_id=tipo.id,
                                classificacao=classificacao_key,
                                estrelas=estrela,
                                preco_por_kg=float(preco_val)
                            )
                            db.session.add(preco_obj)
        
        db.session.commit()
        
        return jsonify(tipo.to_dict()), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao criar tipo de lote: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['PUT'])
@admin_required
def atualizar_tipo_lote(id):
    try:
        tipo = TipoLote.query.get(id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        data = request.get_json()
        
        if not data:
            return jsonify({'erro': 'Dados não fornecidos'}), 400
        
        if data.get('nome') and data['nome'] != tipo.nome:
            tipo_existente = TipoLote.query.filter_by(nome=data['nome']).first()
            if tipo_existente:
                return jsonify({'erro': 'Já existe um tipo de lote com este nome'}), 400
            tipo.nome = data['nome']
        
        if data.get('codigo') and data['codigo'] != tipo.codigo:
            codigo_existente = TipoLote.query.filter_by(codigo=data['codigo']).first()
            if codigo_existente:
                return jsonify({'erro': 'Já existe um tipo de lote com este código'}), 400
            tipo.codigo = data['codigo']
        
        if 'descricao' in data:
            tipo.descricao = data['descricao']
        
        if 'classificacao' in data:
            classificacao = data['classificacao']
            if classificacao == '':
                classificacao = None
            if classificacao and classificacao not in ['leve', 'media', 'pesada']:
                return jsonify({'erro': 'Classificação deve ser: leve, media ou pesada'}), 400
            tipo.classificacao = classificacao
        
        if 'ativo' in data:
            tipo.ativo = data['ativo']
        
        if 'precos' in data and isinstance(data['precos'], dict):
            TipoLotePreco.query.filter_by(tipo_lote_id=tipo.id).delete()
            
            for classificacao_key in ['leve', 'medio', 'pesado']:
                if classificacao_key in data['precos']:
                    precos_class = data['precos'][classificacao_key]
                    for estrela in range(1, 6):
                        preco_val = precos_class.get(str(estrela), 0.0)
                        if isinstance(preco_val, (int, float)) and preco_val >= 0:
                            preco_obj = TipoLotePreco(
                                tipo_lote_id=tipo.id,
                                classificacao=classificacao_key,
                                estrelas=estrela,
                                preco_por_kg=float(preco_val)
                            )
                            db.session.add(preco_obj)
        
        db.session.commit()
        
        return jsonify(tipo.to_dict()), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao atualizar tipo de lote: {str(e)}'}), 500

@bp.route('/<int:id>', methods=['DELETE'])
@admin_required
def deletar_tipo_lote(id):
    try:
        tipo = TipoLote.query.get(id)
        
        if not tipo:
            return jsonify({'erro': 'Tipo de lote não encontrado'}), 404
        
        if len(tipo.itens_solicitacao) > 0 or len(tipo.lotes) > 0:
            return jsonify({'erro': 'Não é possível deletar tipo de lote com solicitações ou lotes associados. Desative-o em vez disso.'}), 400
        
        db.session.delete(tipo)
        db.session.commit()
        
        return jsonify({'mensagem': 'Tipo de lote deletado com sucesso'}), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao deletar tipo de lote: {str(e)}'}), 500

@bp.route('/modelo-importacao', methods=['GET'])
@admin_required
def baixar_modelo_importacao():
    try:
        buffer = criar_modelo_importacao_tipos_lote()
        filename = f'modelo_importacao_tipos_lote_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao gerar modelo: {str(e)}'}), 500

@bp.route('/importar-excel', methods=['POST'])
@admin_required
def importar_excel():
    try:
        if 'arquivo' not in request.files:
            return jsonify({'erro': 'Nenhum arquivo foi enviado'}), 400
        
        arquivo = request.files['arquivo']
        
        if not arquivo.filename or arquivo.filename == '':
            return jsonify({'erro': 'Nenhum arquivo selecionado'}), 400
        
        if not arquivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'erro': 'Formato de arquivo inválido. Use .xlsx ou .xls'}), 400
        
        total_tipos = TipoLote.query.count()
        
        arquivo_bytes = arquivo.read()
        df_tipos = pd.read_excel(io.BytesIO(arquivo_bytes), sheet_name='Tipos de Lote', header=1) if 'Tipos de Lote' in pd.ExcelFile(io.BytesIO(arquivo_bytes)).sheet_names else pd.read_excel(io.BytesIO(arquivo_bytes), header=1)
        
        colunas_requeridas = ['Nome', 'Descrição']
        colunas_faltando = [col for col in colunas_requeridas if col not in df_tipos.columns]
        
        if colunas_faltando:
            return jsonify({'erro': f'Colunas obrigatórias faltando: {", ".join(colunas_faltando)}'}), 400
        
        tipos_criados = 0
        tipos_atualizados = 0
        precos_criados = 0
        erros = []
        
        for idx, (index, row) in enumerate(df_tipos.iterrows()):
            linha_num = idx + 3
            try:
                nome = str(row['Nome']).strip()
                if not nome or nome == 'nan' or pd.isna(row['Nome']):
                    continue
                
                if total_tipos + tipos_criados >= 150:
                    erros.append(f'Linha {linha_num}: Limite máximo de 150 tipos de lote atingido')
                    break
                
                descricao = str(row.get('Descrição', '')).strip()
                if descricao == 'nan' or pd.isna(row.get('Descrição')):
                    descricao = ''
                
                tipo_existente = TipoLote.query.filter_by(nome=nome).first()
                
                if tipo_existente:
                    tipo_existente.descricao = descricao if descricao else tipo_existente.descricao
                    tipo = tipo_existente
                    tipos_atualizados += 1
                    TipoLotePreco.query.filter_by(tipo_lote_id=tipo.id).delete()
                else:
                    codigo = gerar_codigo_automatico()
                    novo_tipo = TipoLote(
                        nome=nome,
                        codigo=codigo,
                        descricao=descricao,
                        ativo=True
                    )
                    db.session.add(novo_tipo)
                    db.session.flush()
                    tipo = novo_tipo
                    tipos_criados += 1
                
                colunas_leve = ['1 Estrela', '2 Estrelas', '3 Estrelas', '4 Estrelas', '5 Estrelas']
                colunas_medio = ['1 Estrela', '2 Estrelas', '3 Estrelas', '4 Estrelas', '5 Estrelas']
                colunas_pesado = ['1 Estrela', '2 Estrelas', '3 Estrelas', '4 Estrelas', '5 Estrelas']
                
                col_offset_leve = 2
                col_offset_medio = 7
                col_offset_pesado = 12
                
                for estrela in range(1, 6):
                    col_idx_leve = col_offset_leve + (estrela - 1)
                    col_idx_medio = col_offset_medio + (estrela - 1)
                    col_idx_pesado = col_offset_pesado + (estrela - 1)
                    
                    try:
                        preco_leve = row.iloc[col_idx_leve] if col_idx_leve < len(row) else None
                        if preco_leve is not None and not pd.isna(preco_leve) and float(preco_leve) > 0:
                            preco_obj = TipoLotePreco(
                                tipo_lote_id=tipo.id,
                                classificacao='leve',
                                estrelas=estrela,
                                preco_por_kg=float(preco_leve)
                            )
                            db.session.add(preco_obj)
                            precos_criados += 1
                    except:
                        pass
                    
                    try:
                        preco_medio = row.iloc[col_idx_medio] if col_idx_medio < len(row) else None
                        if preco_medio is not None and not pd.isna(preco_medio) and float(preco_medio) > 0:
                            preco_obj = TipoLotePreco(
                                tipo_lote_id=tipo.id,
                                classificacao='medio',
                                estrelas=estrela,
                                preco_por_kg=float(preco_medio)
                            )
                            db.session.add(preco_obj)
                            precos_criados += 1
                    except:
                        pass
                    
                    try:
                        preco_pesado = row.iloc[col_idx_pesado] if col_idx_pesado < len(row) else None
                        if preco_pesado is not None and not pd.isna(preco_pesado) and float(preco_pesado) > 0:
                            preco_obj = TipoLotePreco(
                                tipo_lote_id=tipo.id,
                                classificacao='pesado',
                                estrelas=estrela,
                                preco_por_kg=float(preco_pesado)
                            )
                            db.session.add(preco_obj)
                            precos_criados += 1
                    except:
                        pass
            
            except Exception as e:
                erros.append(f'Linha {linha_num}: {str(e)}')
                continue
        
        db.session.commit()
        
        return jsonify({
            'mensagem': 'Importação concluída com sucesso',
            'tipos_criados': tipos_criados,
            'tipos_atualizados': tipos_atualizados,
            'precos_criados': precos_criados,
            'erros': erros
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': f'Erro ao importar arquivo: {str(e)}'}), 500

@bp.route('/exportar-excel', methods=['GET'])
@admin_required
def exportar_excel():
    try:
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        ws_tipos = wb.create_sheet('Tipos de Lote')
        if default_sheet and default_sheet in wb.worksheets:
            wb.remove(default_sheet)
        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        headers_tipos = ['ID', 'Código', 'Nome', 'Classificação', 'Descrição', 'Ativo']
        ws_tipos.append(headers_tipos)
        
        for cell in ws_tipos[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        tipos = TipoLote.query.order_by(TipoLote.nome).all()
        for tipo in tipos:
            ws_tipos.append([
                tipo.id,
                tipo.codigo or '',
                tipo.nome,
                tipo.classificacao,
                tipo.descricao or '',
                'Sim' if tipo.ativo else 'Não'
            ])
        
        for column in ws_tipos.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_tipos.column_dimensions[column[0].column_letter].width = adjusted_width
        
        ws_estrelas = wb.create_sheet('Estrelas por Fornecedor')
        headers_estrelas = ['Tipo de Lote', 'Fornecedor', 'Leve ()', 'Médio ()', 'Pesado ()', 'Ativo']
        ws_estrelas.append(headers_estrelas)
        
        for cell in ws_estrelas[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        classificacoes = FornecedorTipoLoteClassificacao.query.join(
            TipoLote, FornecedorTipoLoteClassificacao.tipo_lote_id == TipoLote.id
        ).join(
            Fornecedor, FornecedorTipoLoteClassificacao.fornecedor_id == Fornecedor.id
        ).order_by(TipoLote.nome, Fornecedor.nome).all()
        
        for classif in classificacoes:
            ws_estrelas.append([
                classif.tipo_lote.nome if classif.tipo_lote else '',
                classif.fornecedor.nome if classif.fornecedor else '',
                classif.leve_estrelas,
                classif.medio_estrelas,
                classif.pesado_estrelas,
                'Sim' if classif.ativo else 'Não'
            ])
        
        for column in ws_estrelas.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_estrelas.column_dimensions[column[0].column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'tipos_lote_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'erro': f'Erro ao exportar Excel: {str(e)}'}), 500

